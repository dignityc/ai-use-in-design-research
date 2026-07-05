#!/usr/bin/env python3
"""Add a discipline-normalized version of the split value heatmap sheet."""

from __future__ import annotations

import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "00_Final.xlsx"
FINAL_SHEET = "LLM1 Validation June 25 (FINAL)"
SOURCE_SHEET = "Heatmap - Split Values"
TARGET_SHEET = "Heatmap - Split Values Norm"
CATEGORIES_SOURCE_SHEET = "Heatmap - Categories"
CATEGORIES_TARGET_SHEET = "Heatmap - Categories Norm"
PHASE_TARGET_SHEET = "Heatmap - Phase Norm"

DISCIPLINES = ("ED", "SD", "UIUX")
VALUE_START_COL = 2
VALUE_END_COL = 43
TOTAL_COL = 44

ENHANCED_ROWS = range(6, 11)
IMPAIRED_ROWS = range(18, 23)

CATEGORY_START_COL = 3
CATEGORY_END_COL = 14
CATEGORY_TOTAL_COL = 15
CATEGORY_DISCIPLINE_ROWS = {
    "ED": range(6, 10),
    "SD": range(10, 14),
    "UIUX": range(14, 18),
}
CATEGORY_SUM_ROW = 18
PHASES = ("Discover", "Define", "Develop", "Deliver")


def blend_hex(start: str, end: str, amount: float) -> str:
    amount = max(0.0, min(1.0, amount))
    s = tuple(int(start[i : i + 2], 16) for i in (0, 2, 4))
    e = tuple(int(end[i : i + 2], 16) for i in (0, 2, 4))
    return "".join(f"{round(sv + (ev - sv) * amount):02X}" for sv, ev in zip(s, e))


def copy_sheet_position(wb, source_name: str, target_name: str):
    if target_name in wb.sheetnames:
        del wb[target_name]
    target = wb.copy_worksheet(wb[source_name])
    target.title = target_name
    return target


def normalize_block(ws, header_row: int, rows: range, color_hex: str) -> dict[str, float]:
    totals = {discipline: 0.0 for discipline in DISCIPLINES}
    for col in range(VALUE_START_COL, VALUE_END_COL + 1):
        discipline = ws.cell(header_row, col).value
        if discipline in totals:
            totals[discipline] += float(ws.cell(rows.stop - 1, col).value or 0)

    normalized_values: list[float] = []
    for row in rows:
        for col in range(VALUE_START_COL, VALUE_END_COL + 1):
            discipline = ws.cell(header_row, col).value
            cell = ws.cell(row, col)
            raw = float(cell.value or 0)
            denom = totals.get(discipline, 0)
            value = raw / denom if denom else 0
            cell.value = value if raw else None
            cell.number_format = "0.0%"
            if raw:
                normalized_values.append(value)

    max_value = max(normalized_values or [1])
    white = "FFFFFF"
    for row in rows:
        for col in range(VALUE_START_COL, VALUE_END_COL + 1):
            cell = ws.cell(row, col)
            value = float(cell.value or 0)
            intensity = (value / max_value) if max_value else 0
            fill = blend_hex(white, color_hex, 0.15 + intensity * 0.75) if value else white
            cell.fill = PatternFill("solid", fgColor=fill)

    ws.cell(rows.stop - 1, TOTAL_COL).value = None
    return totals


def set_note(ws, row: int, totals: dict[str, float], label: str) -> None:
    detail = ", ".join(f"{discipline} n={int(totals[discipline])}" for discipline in DISCIPLINES)
    ws.cell(row, 1).value = (
        f"{label}: counts normalized within each Design Discipline. "
        f"Denominators: {detail}."
    )


def clear_merged_title_count(ws) -> None:
    ws.cell(2, TOTAL_COL).value = None
    ws.cell(14, TOTAL_COL).value = None


def normalize_discipline(value: object) -> str:
    text = str(value or "").replace("\xa0", " ").strip()
    if text == "Product Engineering":
        return "ED"
    if text.startswith("Service"):
        return "SD"
    if text.startswith("UIUX"):
        return "UIUX"
    return ""


def normalize_category_sheet(wb) -> None:
    if CATEGORIES_SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing source sheet: {CATEGORIES_SOURCE_SHEET}")

    ws = copy_sheet_position(wb, CATEGORIES_SOURCE_SHEET, CATEGORIES_TARGET_SHEET)
    ws.sheet_properties.tabColor = "86A8E7"
    ws.cell(1, 1).value = "Category heatmap: normalized share within each Design Discipline"

    totals: dict[str, float] = {}
    for discipline, rows in CATEGORY_DISCIPLINE_ROWS.items():
        total = sum(float(ws.cell(row, CATEGORY_TOTAL_COL).value or 0) for row in rows)
        totals[discipline] = total

    normalized_values: list[float] = []
    for discipline, rows in CATEGORY_DISCIPLINE_ROWS.items():
        denom = totals[discipline]
        for row in rows:
            for col in range(CATEGORY_START_COL, CATEGORY_TOTAL_COL + 1):
                cell = ws.cell(row, col)
                raw = float(cell.value or 0)
                value = raw / denom if denom else 0
                cell.value = value if raw else None
                cell.number_format = "0.0%"
                if raw and col < CATEGORY_TOTAL_COL:
                    normalized_values.append(value)

    total_valid = sum(totals.values())
    for col in range(CATEGORY_START_COL, CATEGORY_TOTAL_COL + 1):
        raw = float(ws.cell(CATEGORY_SUM_ROW, col).value or 0)
        cell = ws.cell(CATEGORY_SUM_ROW, col)
        cell.value = raw / total_valid if raw and total_valid else None
        cell.number_format = "0.0%"

    max_value = max(normalized_values or [1])
    white = "FFFFFF"
    blue = "3867D6"
    for discipline, rows in CATEGORY_DISCIPLINE_ROWS.items():
        for row in rows:
            for col in range(CATEGORY_START_COL, CATEGORY_TOTAL_COL + 1):
                cell = ws.cell(row, col)
                value = float(cell.value or 0)
                intensity = (value / max_value) if max_value and col < CATEGORY_TOTAL_COL else 0
                fill = blend_hex(white, blue, 0.12 + intensity * 0.72) if value and col < CATEGORY_TOTAL_COL else white
                cell.fill = PatternFill("solid", fgColor=fill)

    for col in range(CATEGORY_START_COL, CATEGORY_TOTAL_COL + 1):
        cell = ws.cell(CATEGORY_SUM_ROW, col)
        cell.fill = copy.copy(ws.cell(CATEGORY_SUM_ROW, col).fill)

    detail = ", ".join(f"{discipline} n={int(totals[discipline])}" for discipline in DISCIPLINES)
    ws.cell(2, 1).value = (
        "기준: category counts normalized within each Design Discipline. "
        f"Denominators: {detail}. SUM row uses overall n={int(total_valid)}."
    )


def create_phase_sheet(wb) -> None:
    if FINAL_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing source sheet: {FINAL_SHEET}")

    if PHASE_TARGET_SHEET in wb.sheetnames:
        del wb[PHASE_TARGET_SHEET]

    source = wb[FINAL_SHEET]
    ws = wb.create_sheet(PHASE_TARGET_SHEET)
    ws.sheet_properties.tabColor = "7C3AED"
    ws.freeze_panes = "B5"

    ws.cell(1, 1).value = "Discipline x Design Phase: normalized share within each Design Discipline"
    ws.cell(2, 1).value = (
        "기준: each Design Discipline row sums to 100%. "
        "Cell = phase count / discipline total."
    )

    headers = ("Design discipline", *PHASES, "SUM")
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col)
        cell.value = header
        cell.font = Font(bold=True, color="172033")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="F5F3FF")

    totals: dict[str, float] = {}
    normalized_values: list[float] = []
    raw_counts: dict[str, list[float]] = {}
    for discipline in DISCIPLINES:
        counts = [0.0 for _ in PHASES]
        for row in range(2, source.max_row + 1):
            source_discipline = normalize_discipline(source.cell(row, 2).value)
            phase = str(source.cell(row, 3).value or "").replace("\xa0", " ").strip()
            if source_discipline == discipline and phase in PHASES:
                counts[PHASES.index(phase)] += 1
        total = sum(counts)
        totals[discipline] = total
        raw_counts[discipline] = counts
        normalized_values.extend([(count / total) for count in counts if total and count])

    max_value = max(normalized_values or [1])
    white = "FFFFFF"
    purple = "7C3AED"
    for row_index, discipline in enumerate(DISCIPLINES, 5):
        total = totals[discipline]
        ws.cell(row_index, 1).value = discipline
        ws.cell(row_index, 1).font = Font(bold=True, color="172033")
        ws.cell(row_index, 1).alignment = Alignment(horizontal="center", vertical="center")
        for offset, raw in enumerate(raw_counts[discipline], 2):
            value = raw / total if total else 0
            cell = ws.cell(row_index, offset)
            cell.value = value if raw else None
            cell.number_format = "0.0%"
            cell.alignment = Alignment(horizontal="center", vertical="center")
            intensity = (value / max_value) if max_value else 0
            fill = blend_hex(white, purple, 0.12 + intensity * 0.72) if value else white
            cell.fill = PatternFill("solid", fgColor=fill)
        sum_cell = ws.cell(row_index, len(headers))
        sum_cell.value = sum((raw / total) for raw in raw_counts[discipline]) if total else None
        sum_cell.number_format = "0.0%"
        sum_cell.font = Font(bold=True, color="172033")
        sum_cell.alignment = Alignment(horizontal="center", vertical="center")

    detail = ", ".join(f"{discipline} n={int(totals[discipline])}" for discipline in DISCIPLINES)
    ws.cell(2, 1).value = (
        "기준: phase counts normalized within each Design Discipline. "
        f"Denominators: {detail}."
    )

    ws.column_dimensions["A"].width = 20
    for col in range(2, len(headers) + 1):
        ws.column_dimensions[chr(64 + col)].width = 14
    for row in range(1, 8):
        ws.row_dimensions[row].height = 24


def main() -> None:
    wb = load_workbook(WORKBOOK)
    if SOURCE_SHEET not in wb.sheetnames:
        raise ValueError(f"Missing source sheet: {SOURCE_SHEET}")

    ws = copy_sheet_position(wb, SOURCE_SHEET, TARGET_SHEET)
    ws.sheet_properties.tabColor = "6BAA75"

    ws.cell(1, 1).value = "Enhanced Value heatmap: normalized positive value share"
    ws.cell(13, 1).value = "Impaired Value heatmap: normalized negative value share"

    enhanced_totals = normalize_block(ws, header_row=5, rows=ENHANCED_ROWS, color_hex="1F8A4C")
    impaired_totals = normalize_block(ws, header_row=17, rows=IMPAIRED_ROWS, color_hex="C43131")
    set_note(ws, 2, enhanced_totals, "기준")
    set_note(ws, 14, impaired_totals, "기준")
    clear_merged_title_count(ws)

    normalize_category_sheet(wb)
    create_phase_sheet(wb)

    wb.save(WORKBOOK)


if __name__ == "__main__":
    main()
