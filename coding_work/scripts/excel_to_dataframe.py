"""
Excel 파일에서 셀 값과 코멘트를 읽어 pandas DataFrame으로 변환
CodedPapers_2 시트의 115개 논문 데이터 + 787개 셀 코멘트 추출
"""

import openpyxl
import pandas as pd
from pathlib import Path
from typing import Dict, List, Tuple


class ExcelDataLoader:
    """Excel 파일 로더 (셀 값 + 코멘트)"""

    def __init__(self, excel_path: str, sheet_name: str = "CodedPapers_2"):
        """
        초기화

        Args:
            excel_path: Excel 파일 경로
            sheet_name: 시트 이름 (기본: CodedPapers_2)
        """
        self.excel_path = Path(excel_path)
        self.sheet_name = sheet_name

        # 13개 주요 컬럼 (메타데이터 6개 + 분류 7개)
        self.header_columns = [
            "Index", "Article Title", "Author Full Names", "Source Title",
            "Publication Year", "DOI Link", "Design Discipline", "Data About",
            "Data Modality", "AI methods", "AI Assistance Types (Lee and Kim, 2025)",
            "Design Phase", "Design Practice/Task"
        ]

        # 7개 분류 컬럼 (코멘트가 있는 컬럼)
        self.classification_columns = [
            "Design Discipline",
            "Data About",
            "Data Modality",
            "AI methods",
            "AI Assistance Types (Lee and Kim, 2025)",
            "Design Phase",
            "Design Practice/Task"
        ]

    def load_excel_with_comments(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Excel 파일에서 셀 값과 코멘트 로드

        Returns:
            Tuple[pd.DataFrame, pd.DataFrame]: (셀 값 DataFrame, 코멘트 DataFrame)
            - 셀 값 DataFrame: 13개 컬럼
            - 코멘트 DataFrame: Index + 7개 분류 컬럼 (코멘트 텍스트)
        """
        print(f"📥 Excel 파일 로드 중: {self.excel_path}")

        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 파일을 찾을 수 없습니다: {self.excel_path}")

        # Excel 파일 열기 (read_only=False: 코멘트 읽기 위해 필요)
        wb = openpyxl.load_workbook(self.excel_path, data_only=False)

        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f"시트를 찾을 수 없습니다: {self.sheet_name}")

        ws = wb[self.sheet_name]

        print(f"   ✅ 시트 발견: '{self.sheet_name}'")
        print(f"   📊 총 행 개수: {ws.max_row}")
        print(f"   📊 총 열 개수: {ws.max_column}")

        # 헤더 읽기 (1행)
        headers = []
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(1, col)
            headers.append(cell.value)

        print(f"   📋 헤더: {len(headers)}개 컬럼")

        # 13개 주요 컬럼의 인덱스 찾기
        column_indices = {}
        for col_name in self.header_columns:
            try:
                column_indices[col_name] = headers.index(col_name) + 1  # 1-based index
            except ValueError:
                print(f"   ⚠️  경고: '{col_name}' 컬럼을 찾을 수 없습니다.")
                column_indices[col_name] = None

        # 데이터 읽기 (2행부터)
        values_data = []
        comments_data = []

        for row in range(2, ws.max_row + 1):
            # Index 확인 (빈 행이면 중단)
            index_col = column_indices.get("Index")
            if not index_col:
                break

            index_cell = ws.cell(row, index_col)
            index_value = index_cell.value

            if not index_value:
                # Index가 비어있으면 데이터 종료
                break

            # 셀 값 추출 (13개 컬럼)
            row_values = {"Index": str(index_value).strip()}
            row_comments = {"Index": str(index_value).strip()}

            for col_name, col_idx in column_indices.items():
                if col_name == "Index" or not col_idx:
                    continue

                cell = ws.cell(row, col_idx)
                cell_value = cell.value if cell.value else ""

                # 셀 값 저장
                row_values[col_name] = str(cell_value).strip()

                # 7개 분류 컬럼은 코멘트도 저장
                if col_name in self.classification_columns:
                    if cell.comment:
                        comment_text = cell.comment.text
                        row_comments[col_name] = comment_text.strip()
                    else:
                        row_comments[col_name] = ""

            values_data.append(row_values)
            comments_data.append(row_comments)

        wb.close()

        print(f"   ✅ 데이터 로드 완료: {len(values_data)}개 논문")

        # DataFrame 생성
        values_df = pd.DataFrame(values_data)
        comments_df = pd.DataFrame(comments_data)

        # 컬럼 순서 보장
        values_df = values_df[self.header_columns]
        comments_df = comments_df[["Index"] + self.classification_columns]

        # Index 유니크성 검증
        duplicate_indices = values_df[values_df.duplicated(subset=['Index'], keep=False)]
        if not duplicate_indices.empty:
            print(f"   ⚠️  경고: 중복된 Index 발견!")
            print(duplicate_indices[['Index', 'Article Title']])
        else:
            print(f"   ✅ Index 유니크성 검증: 모든 Index가 유니크함")

        # 코멘트 통계
        total_comments = 0
        for col in self.classification_columns:
            col_comments = comments_df[col].apply(lambda x: bool(x and x.strip()))
            count = col_comments.sum()
            total_comments += count
            print(f"   📝 {col}: {count}개 코멘트")

        print(f"   📝 총 셀 코멘트: {total_comments}개")

        return values_df, comments_df


def main():
    """테스트 실행"""
    # Excel 파일 경로
    excel_path = Path(__file__).parent.parent / "Coding & screening works-integrated.xlsx"

    loader = ExcelDataLoader(excel_path)

    try:
        values_df, comments_df = loader.load_excel_with_comments()

        print("\n" + "=" * 60)
        print("📊 셀 값 DataFrame")
        print("=" * 60)
        print(f"행 개수: {len(values_df)}")
        print(f"열 개수: {len(values_df.columns)}")
        print("\n첫 3개 논문:")
        print(values_df.head(3)[['Index', 'Article Title', 'Design Discipline', 'AI methods']])

        print("\n" + "=" * 60)
        print("📝 코멘트 DataFrame")
        print("=" * 60)
        print(f"행 개수: {len(comments_df)}")
        print(f"열 개수: {len(comments_df.columns)}")
        print("\n첫 3개 논문:")
        print(comments_df.head(3)[['Index', 'Design Discipline', 'AI methods']])

        return True

    except Exception as e:
        print(f"\n❌ 오류 발생: {e}")
        import traceback
        traceback.print_exc()
        return False


if __name__ == "__main__":
    main()
